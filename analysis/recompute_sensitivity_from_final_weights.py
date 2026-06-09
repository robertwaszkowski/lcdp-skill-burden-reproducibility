#!/usr/bin/env python3
from pathlib import Path
import itertools
import math
import sys

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PKG = ROOT / "data" / "platform_score_package"
DATASET = PKG / "lcdp_skill_burden_dataset.xlsx"
CALC = PKG / "lcdp_skill_burden_calculations.xlsx"

PLATFORMS = [
    "Aurea",
    "Google AppSheet",
    "OutSystems",
    "Zoho Creator",
    "Microsoft Power Apps",
    "Mendix",
]

PHASES = ["Design", "Development", "Implementation"]

APPLICATION_LOGIC_TERMS = [
    "logic",
    "program",
    "javascript",
    "java",
    "typescript",
    ".net",
    "function",
    "formula",
    "algorithm",
    "variable",
    "script",
    "xml",
    "xpath",
    "sql",
]


def norm_skill_id(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.upper().startswith("SK"):
        digits = "".join(ch for ch in s if ch.isdigit())
        return f"SK{int(digits):03d}" if digits else s
    if s.replace(".", "", 1).isdigit():
        return f"SK{int(float(s)):03d}"
    return s


def load_data():
    matrix = pd.read_excel(DATASET, sheet_name="Platform_Skill_Matrix")
    weights = pd.read_excel(DATASET, sheet_name="Skill_Taxonomy_Weights")

    matrix = matrix.copy()
    weights = weights.copy()

    matrix["Skill_ID_norm"] = matrix["Skill_ID"].map(norm_skill_id)
    weights["Skill_ID_norm"] = weights["Skill_ID"].map(norm_skill_id)

    for p in PLATFORMS:
        matrix[p] = matrix[p].fillna(0).astype(float)

    df = matrix[matrix["Skill_ID_norm"] != ""].merge(
        weights[[
            "Skill_ID_norm",
            "Skill",
            "Project_Stage",
            "Project_Task",
            "Usage_Intensity",
            "Difficulty_of_Acquisition",
            "Synthetic_Weight",
        ]],
        on="Skill_ID_norm",
        how="left",
        suffixes=("", "_weight"),
        validate="many_to_one",
    )

    if df["Synthetic_Weight"].isna().any():
        bad = df[df["Synthetic_Weight"].isna()]
        print(bad[["Skill_ID", "Project_Stage", "Project_Task", "Skill"]].head(20).to_string(index=False))
        raise SystemExit("ERROR: missing weights after merge.")

    return matrix, weights, df


def compute_phase_scores(df, weight_multiplier=None, p=1):
    work = df.copy()

    if weight_multiplier is None:
        weight_multiplier = pd.Series(1.0, index=work.index)
    elif not isinstance(weight_multiplier, pd.Series):
        weight_multiplier = pd.Series(weight_multiplier, index=work.index)

    work["Effective_Weight"] = work["Synthetic_Weight"] * weight_multiplier

    rows = []
    for platform in PLATFORMS:
        phase_values = {}
        for phase in PHASES:
            sub = work[work["Project_Stage"] == phase]
            if p == 1:
                value = (sub[platform] * sub["Effective_Weight"]).sum()
            else:
                value = ((sub[platform] * sub["Effective_Weight"]) ** p).sum() ** (1.0 / p)
            phase_values[phase] = float(value)

        total = sum(phase_values.values())
        rows.append({
            "Platform": platform,
            "Design": phase_values["Design"],
            "Development": phase_values["Development"],
            "Implementation": phase_values["Implementation"],
            "Total_Score": total,
        })

    out = pd.DataFrame(rows)
    out["Rank"] = out["Total_Score"].rank(method="min", ascending=True).astype(int)
    return out.sort_values("Rank").reset_index(drop=True)


def rank_from_phase_scores(phase_scores, omega=None, q=1):
    if omega is None:
        omega = {"Design": 1.0, "Development": 1.0, "Implementation": 1.0}

    rows = []
    for _, row in phase_scores.iterrows():
        vals = [float(row[ph]) for ph in PHASES]
        weights = [float(omega[ph]) for ph in PHASES]

        if math.isinf(q):
            total = max(w * v for w, v in zip(weights, vals))
        elif q == 1:
            total = sum(w * v for w, v in zip(weights, vals))
        else:
            total = sum(w * (v ** q) for w, v in zip(weights, vals)) ** (1.0 / q)

        rows.append({
            "Platform": row["Platform"],
            "Total_Score": total,
        })

    out = pd.DataFrame(rows)
    out["Rank"] = out["Total_Score"].rank(method="min", ascending=True).astype(int)
    return out.sort_values("Rank").reset_index(drop=True)


def scenario_rank_table(base_ranking, scenarios):
    rows = []
    for platform in PLATFORMS:
        base_rank = int(base_ranking.loc[base_ranking["Platform"] == platform, "Rank"].iloc[0])
        row = {"Platform": platform, "Base_Rank": base_rank}
        for label, ranking in scenarios.items():
            rank = int(ranking.loc[ranking["Platform"] == platform, "Rank"].iloc[0])
            row[f"{label}_Rank"] = rank
            row[f"{label}_Change"] = rank - base_rank
        rows.append(row)
    return pd.DataFrame(rows).sort_values("Base_Rank").reset_index(drop=True)


def spearman_from_ranks(base_ranking, scenario_ranking):
    a = base_ranking.set_index("Platform").loc[PLATFORMS, "Rank"].astype(float)
    b = scenario_ranking.set_index("Platform").loc[PLATFORMS, "Rank"].astype(float)
    return float(a.corr(b, method="spearman"))


def detect_application_logic(df):
    text = (
        df["Project_Task"].fillna("").astype(str).str.lower()
        + " "
        + df["Skill"].fillna("").astype(str).str.lower()
    )
    mask = False
    for term in APPLICATION_LOGIC_TERMS:
        mask = mask | text.str.contains(term, regex=False)
    return mask


def first_rank_change_threshold(df, mask, direction, max_pct=2.0, step=0.01):
    base = compute_phase_scores(df)
    base_order = tuple(base.sort_values(["Rank", "Platform"])["Platform"])
    pct = step
    while pct <= max_pct + 1e-12:
        mult = pd.Series(1.0, index=df.index)
        if direction == "increase":
            mult.loc[mask] = 1.0 + pct
        else:
            mult.loc[mask] = max(0.0, 1.0 - pct)
        rank = compute_phase_scores(df, weight_multiplier=mult)
        order = tuple(rank.sort_values(["Rank", "Platform"])["Platform"])
        if order != base_order:
            return pct
        pct += step
    return None


def switching_points(base_phase):
    base = base_phase.sort_values("Rank").reset_index(drop=True)
    avg_weight = float(pd.read_excel(DATASET, sheet_name="Skill_Taxonomy_Weights")["Synthetic_Weight"].mean())

    rows = []
    for i in range(len(base) - 1):
        a = base.iloc[i]
        b = base.iloc[i + 1]

        diffs = {
            "Design": float(b["Design"] - a["Design"]),
            "Development": float(b["Development"] - a["Development"]),
            "Implementation": float(b["Implementation"] - a["Implementation"]),
        }
        total_gap = float(b["Total_Score"] - a["Total_Score"])

        # Boundary under omega_A + omega_D + omega_I = 1:
        # diff_A*omega_A + diff_D*omega_D + diff_I*(1-omega_A-omega_D)=0
        # (diff_A - diff_I) omega_A + (diff_D - diff_I) omega_D = -diff_I
        coef_a = diffs["Design"] - diffs["Implementation"]
        coef_d = diffs["Development"] - diffs["Implementation"]
        rhs = -diffs["Implementation"]

        dominant_phase = max(diffs, key=lambda ph: abs(diffs[ph]))
        resilience = math.ceil(total_gap / avg_weight) if avg_weight > 0 else None

        rows.append({
            "Adjacent_Pair": f"{a['Platform']} vs. {b['Platform']}",
            "Total_Gap": total_gap,
            "Dominant_Phase_Difference": dominant_phase,
            "Delta_Design": diffs["Design"],
            "Delta_Development": diffs["Development"],
            "Delta_Implementation": diffs["Implementation"],
            "Boundary_Coeff_Design": coef_a,
            "Boundary_Coeff_Development": coef_d,
            "Boundary_RHS": rhs,
            "Boundary_Equation": f"{coef_a:.2f} omega_A + {coef_d:.2f} omega_D = {rhs:.2f}",
            "Average_Skill_Weight": avg_weight,
            "Resilience_Errors_Ceil": int(resilience),
        })

    return pd.DataFrame(rows)


def write_sheets(calc_path, sheets):
    wb = load_workbook(calc_path)

    for sheet_name in sheets:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    wb.save(calc_path)

    with pd.ExcelWriter(calc_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def main():
    matrix, weights, df = load_data()

    base_phase = compute_phase_scores(df)
    base_rank = rank_from_phase_scores(base_phase)

    app_logic_mask = detect_application_logic(df)

    plus20_mult = pd.Series(1.0, index=df.index)
    plus20_mult.loc[app_logic_mask] = 1.2
    minus20_mult = pd.Series(1.0, index=df.index)
    minus20_mult.loc[app_logic_mask] = 0.8

    plus20_rank = compute_phase_scores(df, weight_multiplier=plus20_mult)
    minus20_rank = compute_phase_scores(df, weight_multiplier=minus20_mult)

    skill_sens = scenario_rank_table(
        base_phase,
        {
            "ApplicationLogic_plus20": plus20_rank,
            "ApplicationLogic_minus20": minus20_rank,
        },
    )

    plus_threshold = first_rank_change_threshold(df, app_logic_mask, "increase")
    minus_threshold = first_rank_change_threshold(df, app_logic_mask, "decrease")

    skill_summary = pd.DataFrame([{
        "Application_Logic_Skill_Rows": int(app_logic_mask.sum()),
        "Spearman_plus20": spearman_from_ranks(base_phase, plus20_rank),
        "Spearman_minus20": spearman_from_ranks(base_phase, minus20_rank),
        "First_Rank_Change_Increase_Pct_Approx": plus_threshold,
        "First_Rank_Change_Decrease_Pct_Approx": minus_threshold,
    }])

    phase_scenarios = {
        "Equal_Normalized": rank_from_phase_scores(base_phase, {
            "Design": 1/3, "Development": 1/3, "Implementation": 1/3
        }),
        "Development_060": rank_from_phase_scores(base_phase, {
            "Design": 0.2, "Development": 0.6, "Implementation": 0.2
        }),
        "Implementation_Deemphasized_010": rank_from_phase_scores(base_phase, {
            "Design": 0.45, "Development": 0.45, "Implementation": 0.10
        }),
        "Development_090": rank_from_phase_scores(base_phase, {
            "Design": 0.05, "Development": 0.90, "Implementation": 0.05
        }),
    }
    phase_sens = scenario_rank_table(base_rank, phase_scenarios)

    p2_phase = compute_phase_scores(df, p=2)
    p2_q1_rank = rank_from_phase_scores(p2_phase, q=1)
    qinf_rank = rank_from_phase_scores(base_phase, q=math.inf)

    agg_sens = scenario_rank_table(base_rank, {
        "p2_q1": p2_q1_rank,
        "p1_qinf": qinf_rank,
    })

    switching = switching_points(base_phase)

    output_dir = PKG / "final15_sensitivity_outputs"
    output_dir.mkdir(parents=True, exist_ok=True)

    base_phase.to_csv(output_dir / "phase_scores_final15.csv", index=False)
    skill_sens.to_csv(output_dir / "sensitivity_skill_weights_final15.csv", index=False)
    skill_summary.to_csv(output_dir / "sensitivity_skill_weights_summary_final15.csv", index=False)
    phase_sens.to_csv(output_dir / "sensitivity_phase_weights_final15.csv", index=False)
    agg_sens.to_csv(output_dir / "sensitivity_aggregation_final15.csv", index=False)
    switching.to_csv(output_dir / "switching_points_final15.csv", index=False)

    write_sheets(CALC, {
        "Sens_Skill_Final15": skill_sens,
        "Sens_Skill_Summary15": skill_summary,
        "Sens_Phase_Final15": phase_sens,
        "Sens_Aggregation15": agg_sens,
        "Switching_Final15": switching,
    })

    print("=== Final 15-expert base phase scores ===")
    print(base_phase.to_string(index=False))

    print()
    print("=== Skill-weight sensitivity ===")
    print(skill_sens.to_string(index=False))
    print()
    print(skill_summary.to_string(index=False))

    print()
    print("=== Phase-weight sensitivity ===")
    print(phase_sens.to_string(index=False))

    print()
    print("=== Aggregation sensitivity ===")
    print(agg_sens.to_string(index=False))

    print()
    print("=== Switching points ===")
    print(switching.to_string(index=False))

    print()
    print("Outputs written to:")
    print(output_dir)
    print("Updated workbook:")
    print(CALC)

    return 0


if __name__ == "__main__":
    sys.exit(main())
