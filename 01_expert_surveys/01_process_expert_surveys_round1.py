#!/usr/bin/env python3
"""
Reproduce the expert Round 1 survey analysis workbook.

This script reads a ZIP archive containing 15 completed expert survey Excel files,
extracts the ratings, validates the files, computes consensus weights,
expert-response similarity, and reliability statistics, then writes:

  1. expert_survey_round1_analysis_report_round1.xlsx
  2. CSV files with long-format data, consensus weights, reliability metrics,
     validation summary, rating distributions, and pairwise similarity.

Usage example:

  python reproduce_expert_survey_round1_analysis.py \
      --zip completed_expert_surveys_round1_a\\(1\\).zip \
      --output-report expert_survey_round1_analysis_report_round1.xlsx \
      --output-dir expert_analysis_outputs_round1 \
      --output-zip expert_survey_round1_analysis_outputs_round1.zip

Dependencies:
  pip install openpyxl numpy
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import statistics
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


USAGE_MAP = {
    "1 - Rare / avoidable": 1,
    "2 - Occasional / peripheral": 2,
    "3 - Moderate / recurring": 3,
    "4 - Frequent / important": 4,
    "5 - Essential / continuous": 5,
}

DIFFICULTY_MAP = {
    "1 - Very easy": 1,
    "2 - Easy": 2,
    "3 - Moderate": 3,
    "4 - Difficult": 4,
    "5 - Very difficult": 5,
}


def parse_rating(value, mapping: dict[str, int]) -> int | None:
    """Convert a dropdown label, a numeric value, or a label without number to 1..5."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if int(value) == float(value) and 1 <= int(value) <= 5:
            return int(value)
        return None

    text = str(value).strip()
    if text in mapping:
        return mapping[text]

    prefix = re.match(r"^([1-5])\b", text)
    if prefix:
        return int(prefix.group(1))

    labels = {
        key.split(" - ", 1)[1].lower(): score
        for key, score in mapping.items()
        if " - " in key
    }
    return labels.get(text.lower())


def mean(values):
    values = [v for v in values if v is not None]
    return sum(values) / len(values) if values else None


def sample_sd(values):
    values = [v for v in values if v is not None]
    return statistics.stdev(values) if len(values) > 1 else 0


def median(values):
    values = [v for v in values if v is not None]
    return statistics.median(values) if values else None


def pivot_matrix(rows: list[dict], field: str):
    expert_ids = sorted({row["expert_id"] for row in rows})
    skill_ids = sorted({row["skill_id"] for row in rows})
    values = {(row["skill_id"], row["expert_id"]): row[field] for row in rows}
    matrix = [[values.get((skill_id, expert_id)) for expert_id in expert_ids] for skill_id in skill_ids]
    return skill_ids, expert_ids, matrix


def cronbach_alpha(matrix: list[list[float]]) -> float | None:
    """Cronbach's alpha treating skills as cases and experts as raters/items."""
    if not matrix or len(matrix[0]) < 2 or len(matrix) < 2:
        return None
    n_items = len(matrix[0])
    columns = list(zip(*matrix))
    item_variances = [statistics.variance([x for x in col if x is not None]) for col in columns]
    totals = [sum(row) for row in matrix]
    total_variance = statistics.variance(totals)
    if total_variance == 0:
        return None
    return n_items / (n_items - 1) * (1 - sum(item_variances) / total_variance)


def icc_2_1_and_2_k(matrix: list[list[float]]):
    """
    Two-way random-effects absolute-agreement ICC.

    Returns ICC(2,1), ICC(2,k), MS_rows, MS_columns, MS_error.
    Rows are skills, columns are experts.
    """
    x = np.array(matrix, dtype=float)
    n, k = x.shape
    grand_mean = x.mean()
    row_means = x.mean(axis=1)
    col_means = x.mean(axis=0)

    ss_rows = k * ((row_means - grand_mean) ** 2).sum()
    ss_cols = n * ((col_means - grand_mean) ** 2).sum()
    ss_error = ((x - row_means[:, None] - col_means[None, :] + grand_mean) ** 2).sum()

    ms_rows = ss_rows / (n - 1)
    ms_cols = ss_cols / (k - 1)
    ms_error = ss_error / ((n - 1) * (k - 1))

    denominator_single = ms_rows + (k - 1) * ms_error + k * (ms_cols - ms_error) / n
    denominator_average = ms_rows + (ms_cols - ms_error) / n

    icc_single = (ms_rows - ms_error) / denominator_single if denominator_single != 0 else None
    icc_average = (ms_rows - ms_error) / denominator_average if denominator_average != 0 else None
    return icc_single, icc_average, ms_rows, ms_cols, ms_error


def write_csv(path: Path, rows: list[dict]):
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def extract_zip(zip_path: Path, extract_dir: Path):
    if extract_dir.exists():
        shutil.rmtree(extract_dir)
    extract_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path) as archive:
        archive.extractall(extract_dir)


def read_surveys(extract_dir: Path):
    files = sorted([path for path in extract_dir.rglob("*.xlsx") if not path.name.startswith("~$")])
    validations = []
    long_rows = []

    for file_path in files:
        workbook = load_workbook(file_path, data_only=False, read_only=False)
        if "Survey" not in workbook.sheetnames:
            validations.append({"file_name": file_path.name, "status": "FAIL", "issue": "No Survey sheet"})
            continue

        sheet = workbook["Survey"]
        expert_id = sheet["A2"].value
        round_no = sheet["B2"].value
        years_band = sheet["D2"].value
        lcdp_count = sheet["E2"].value
        missing_formula = sheet["G2"].value
        status_formula = sheet["H2"].value

        row_count = 0
        missing = 0
        invalid = 0
        invalid_detail = []
        skill_ids = []

        # Survey rows are expected in rows 5..213 inclusive.
        for row_number in range(5, 214):
            project_stage = sheet.cell(row_number, 1).value
            project_task = sheet.cell(row_number, 2).value
            skill_id = sheet.cell(row_number, 3).value
            skill_name = sheet.cell(row_number, 4).value
            usage_raw = sheet.cell(row_number, 5).value
            difficulty_raw = sheet.cell(row_number, 6).value
            comment = sheet.cell(row_number, 7).value if sheet.max_column >= 7 else None

            if all(v is None for v in [project_stage, project_task, skill_id, skill_name, usage_raw, difficulty_raw]):
                continue

            row_count += 1
            skill_ids.append(skill_id)
            usage = parse_rating(usage_raw, USAGE_MAP)
            difficulty = parse_rating(difficulty_raw, DIFFICULTY_MAP)

            if usage is None:
                if usage_raw is None or str(usage_raw).strip() == "":
                    missing += 1
                else:
                    invalid += 1
                    invalid_detail.append(f"E{row_number}={usage_raw}")
            if difficulty is None:
                if difficulty_raw is None or str(difficulty_raw).strip() == "":
                    missing += 1
                else:
                    invalid += 1
                    invalid_detail.append(f"F{row_number}={difficulty_raw}")

            if usage is not None and difficulty is not None:
                long_rows.append({
                    "file_name": file_path.name,
                    "expert_id": expert_id,
                    "round": round_no,
                    "years_experience_band": years_band,
                    "lcdp_experience_count": lcdp_count,
                    "project_stage": project_stage,
                    "project_task": project_task,
                    "skill_id": int(skill_id) if isinstance(skill_id, (int, float)) else skill_id,
                    "skill_name": skill_name,
                    "usage_intensity_label": usage_raw,
                    "usage_intensity": usage,
                    "difficulty_acquisition_label": difficulty_raw,
                    "difficulty_acquisition": difficulty,
                    "synthetic_weight": usage * difficulty,
                    "comment": comment,
                })

        issue = ""
        if not expert_id:
            issue += "Missing expert_id; "
        if round_no != 1:
            issue += f"Round not 1 ({round_no}); "
        if row_count != 209:
            issue += f"Expected 209 rows, got {row_count}; "
        if sorted(skill_ids) != list(range(1, 210)):
            issue += "Skill IDs are not 1..209; "

        status = "PASS" if not issue and missing == 0 and invalid == 0 else "FAIL"
        validations.append({
            "file_name": file_path.name,
            "expert_id": expert_id,
            "round": round_no,
            "years_experience_band": years_band,
            "lcdp_experience_count": lcdp_count,
            "skill_rows": row_count,
            "missing_rating_cells": missing,
            "invalid_rating_cells": invalid,
            "status": status,
            "issue": issue + "; ".join(invalid_detail[:10]),
            "missing_formula": missing_formula,
            "status_formula": status_formula,
        })

    return files, validations, long_rows


def compute_outputs(long_rows: list[dict]):
    reliability = []
    for label, field in [
        ("Usage Intensity", "usage_intensity"),
        ("Difficulty of Acquisition", "difficulty_acquisition"),
        ("Synthetic Weight", "synthetic_weight"),
    ]:
        skill_ids, expert_ids, matrix = pivot_matrix(long_rows, field)
        alpha = cronbach_alpha(matrix)
        icc_single, icc_average, ms_rows, ms_cols, ms_error = icc_2_1_and_2_k(matrix)
        flat = [value for row in matrix for value in row]
        reliability.append({
            "measure": label,
            "n_skills": len(skill_ids),
            "n_experts": len(expert_ids),
            "mean": mean(flat),
            "sd": sample_sd(flat),
            "min": min(flat),
            "max": max(flat),
            "cronbach_alpha": alpha,
            "icc_2_1_single_expert": icc_single,
            "icc_2_k_average_experts": icc_average,
            "ms_rows": ms_rows,
            "ms_cols": ms_cols,
            "ms_error": ms_error,
        })

    by_skill = defaultdict(list)
    for row in long_rows:
        by_skill[row["skill_id"]].append(row)

    consensus = []
    for skill_id, rows in sorted(by_skill.items()):
        usage_values = [row["usage_intensity"] for row in rows]
        difficulty_values = [row["difficulty_acquisition"] for row in rows]
        weight_values = [row["synthetic_weight"] for row in rows]
        first = rows[0]
        consensus.append({
            "skill_id": skill_id,
            "project_stage": first["project_stage"],
            "project_task": first["project_task"],
            "skill_name": first["skill_name"],
            "n_experts": len(rows),
            "usage_mean": mean(usage_values),
            "usage_median": median(usage_values),
            "usage_sd": sample_sd(usage_values),
            "usage_min": min(usage_values),
            "usage_max": max(usage_values),
            "usage_range": max(usage_values) - min(usage_values),
            "difficulty_mean": mean(difficulty_values),
            "difficulty_median": median(difficulty_values),
            "difficulty_sd": sample_sd(difficulty_values),
            "difficulty_min": min(difficulty_values),
            "difficulty_max": max(difficulty_values),
            "difficulty_range": max(difficulty_values) - min(difficulty_values),
            "weight_mean": mean(weight_values),
            "weight_median": median(weight_values),
            "weight_sd": sample_sd(weight_values),
            "weight_min": min(weight_values),
            "weight_max": max(weight_values),
            "weight_range": max(weight_values) - min(weight_values),
            "consensus_flag": "High" if sample_sd(weight_values) <= 3 else ("Moderate" if sample_sd(weight_values) <= 5 else "Low"),
        })

    usage_counts = Counter(row["usage_intensity"] for row in long_rows)
    difficulty_counts = Counter(row["difficulty_acquisition"] for row in long_rows)
    weight_counts = Counter(row["synthetic_weight"] for row in long_rows)
    distributions = []
    n = len(long_rows)
    for value in range(1, 6):
        distributions.append({"type": "usage_intensity", "value": value, "count": usage_counts[value], "percent": usage_counts[value] / n})
    for value in range(1, 6):
        distributions.append({"type": "difficulty_acquisition", "value": value, "count": difficulty_counts[value], "percent": difficulty_counts[value] / n})
    for value in sorted(weight_counts):
        distributions.append({"type": "synthetic_weight", "value": value, "count": weight_counts[value], "percent": weight_counts[value] / n})

    expert_ids = sorted({row["expert_id"] for row in long_rows})
    expert_values = {expert_id: {} for expert_id in expert_ids}
    for row in long_rows:
        expert_values[row["expert_id"]][row["skill_id"]] = (
            row["usage_intensity"],
            row["difficulty_acquisition"],
            row["synthetic_weight"],
        )

    similarity = []
    similarity_matrix = []
    for expert_1 in expert_ids:
        matrix_row = []
        for expert_2 in expert_ids:
            n_skills = 0
            both_same = 0
            usage_same = 0
            difficulty_same = 0
            weight_same = 0
            cell_count = 0
            cells_same = 0
            for skill_id in sorted(by_skill):
                value_1 = expert_values[expert_1].get(skill_id)
                value_2 = expert_values[expert_2].get(skill_id)
                if value_1 and value_2:
                    n_skills += 1
                    both_same += int(value_1[0] == value_2[0] and value_1[1] == value_2[1])
                    usage_same += int(value_1[0] == value_2[0])
                    difficulty_same += int(value_1[1] == value_2[1])
                    weight_same += int(value_1[2] == value_2[2])
                    cell_count += 2
                    cells_same += int(value_1[0] == value_2[0]) + int(value_1[1] == value_2[1])
            similarity_ratio = both_same / n_skills if n_skills else None
            matrix_row.append(similarity_ratio)
            if expert_1 < expert_2:
                similarity.append({
                    "expert_1": expert_1,
                    "expert_2": expert_2,
                    "n_skills_compared": n_skills,
                    "both_usage_and_difficulty_same_count": both_same,
                    "both_usage_and_difficulty_same_percent": similarity_ratio,
                    "usage_same_percent": usage_same / n_skills if n_skills else None,
                    "difficulty_same_percent": difficulty_same / n_skills if n_skills else None,
                    "cell_level_same_percent": cells_same / cell_count if cell_count else None,
                    "synthetic_weight_same_percent": weight_same / n_skills if n_skills else None,
                    "duplicate_flag": "YES" if both_same == n_skills else ("HIGH_SIMILARITY" if similarity_ratio is not None and similarity_ratio >= 0.95 else ""),
                })
        similarity_matrix.append(matrix_row)

    return reliability, consensus, distributions, similarity, expert_ids, similarity_matrix


def add_sheet(workbook: Workbook, name: str, rows: list[dict], percent_cols: set[str] | None = None):
    sheet = workbook.create_sheet(name)
    if not rows:
        return sheet

    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"

    percent_cols = percent_cols or set()
    for col_idx, header in enumerate(headers, start=1):
        preview_rows = range(2, min(sheet.max_row, 200) + 1)
        max_len = len(str(header))
        for row_idx in preview_rows:
            value = sheet.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 55))
        if header in percent_cols:
            for row_idx in range(2, sheet.max_row + 1):
                sheet.cell(row_idx, col_idx).number_format = "0.0%"
        elif any(key in header for key in ["mean", "sd", "alpha", "icc", "percent"]):
            for row_idx in range(2, sheet.max_row + 1):
                sheet.cell(row_idx, col_idx).number_format = "0.0000"

    ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    try:
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", name)[:25] + "Table"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    except Exception:
        pass
    return sheet


def write_report(
    report_path: Path,
    source_zip: Path,
    files: list[Path],
    validations: list[dict],
    long_rows: list[dict],
    reliability: list[dict],
    consensus: list[dict],
    distributions: list[dict],
    similarity: list[dict],
    expert_ids: list[str],
    similarity_matrix: list[list[float]],
):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    summary_rows = [
        {"item": "Source ZIP", "value": source_zip.name},
        {"item": "Excel files found", "value": len(files)},
        {"item": "Unique expert IDs", "value": len(expert_ids)},
        {"item": "Expected observations", "value": 15 * 209},
        {"item": "Extracted valid observations", "value": len(long_rows)},
        {"item": "Missing rating cells", "value": sum(row.get("missing_rating_cells", 0) for row in validations)},
        {"item": "Invalid rating cells", "value": sum(row.get("invalid_rating_cells", 0) for row in validations)},
        {"item": "Duplicate pairs", "value": sum(1 for row in similarity if row["duplicate_flag"] == "YES")},
        {"item": "High-similarity pairs (>=95%)", "value": sum(1 for row in similarity if row["duplicate_flag"] == "HIGH_SIMILARITY")},
    ]

    summary.append(["item", "value"])
    for row in summary_rows:
        summary.append([row["item"], row["value"]])
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    summary.column_dimensions["A"].width = 35
    summary.column_dimensions["B"].width = 30

    add_sheet(workbook, "FileValidation", validations)
    add_sheet(workbook, "Reliability", reliability)
    add_sheet(
        workbook,
        "PairwiseSimilarity",
        similarity,
        percent_cols={
            "both_usage_and_difficulty_same_percent",
            "usage_same_percent",
            "difficulty_same_percent",
            "cell_level_same_percent",
            "synthetic_weight_same_percent",
        },
    )
    add_sheet(workbook, "ConsensusBySkill", consensus)
    add_sheet(workbook, "Distributions", distributions, percent_cols={"percent"})

    matrix_sheet = workbook.create_sheet("SimilarityMatrix")
    matrix_sheet.append(["Expert ID"] + expert_ids)
    for expert_id, row in zip(expert_ids, similarity_matrix):
        matrix_sheet.append([expert_id] + row)
    for cell in matrix_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_idx in range(2, matrix_sheet.max_row + 1):
        matrix_sheet.cell(row_idx, 1).fill = PatternFill("solid", fgColor="D9EAF7")
        for col_idx in range(2, matrix_sheet.max_column + 1):
            matrix_sheet.cell(row_idx, col_idx).number_format = "0.0%"
    matrix_sheet.freeze_panes = "B2"
    for col_idx in range(1, matrix_sheet.max_column + 1):
        matrix_sheet.column_dimensions[get_column_letter(col_idx)].width = 12
    matrix_sheet.conditional_formatting.add(
        f"B2:{get_column_letter(matrix_sheet.max_column)}{matrix_sheet.max_row}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="FFFFFF",
            mid_type="num", mid_value=0.5, mid_color="FFE699",
            end_type="num", end_value=1, end_color="C6E0B4",
        ),
    )

    add_sheet(workbook, "LongData", long_rows)
    workbook.save(report_path)


def main():
    parser = argparse.ArgumentParser(description="Analyze completed Round 1 expert survey files.")
    parser.add_argument("--zip", default=Path(__file__).parent / "input" / "completed_expert_surveys_round1_FINAL.zip", type=Path, help="ZIP archive containing completed expert survey XLSX files.")
    parser.add_argument("--output-report", default=Path(__file__).parent / "output" / "expert_survey_round1_analysis_report.xlsx", type=Path, help="Output XLSX report path.")
    parser.add_argument("--output-dir", default=Path(__file__).parent / "output", type=Path, help="Directory for CSV outputs.")
    parser.add_argument("--output-zip", default=Path(__file__).parent / "output" / "expert_survey_round1_outputs.zip", type=Path, help="ZIP archive containing report and CSV outputs.")
    parser.add_argument("--extract-dir", default=Path(__file__).parent / "input" / "_extract", type=Path, help="Temporary extraction directory.")
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    for old_file in args.output_dir.glob("*"):
        if old_file.is_file():
            old_file.unlink()

    extract_zip(args.zip, args.extract_dir)
    files, validations, long_rows = read_surveys(args.extract_dir)
    reliability, consensus, distributions, similarity, expert_ids, similarity_matrix = compute_outputs(long_rows)

    write_csv(args.output_dir / "expert_survey_validation_summary.csv", validations)
    write_csv(args.output_dir / "expert_survey_round1_long_format.csv", long_rows)
    write_csv(args.output_dir / "expert_consensus_weights_round1.csv", consensus)
    write_csv(args.output_dir / "expert_reliability_round1.csv", reliability)
    write_csv(args.output_dir / "expert_pairwise_similarity_round1.csv", similarity)
    write_csv(args.output_dir / "expert_rating_distributions_round1.csv", distributions)

    write_report(
        report_path=args.output_report,
        source_zip=args.zip,
        files=files,
        validations=validations,
        long_rows=long_rows,
        reliability=reliability,
        consensus=consensus,
        distributions=distributions,
        similarity=similarity,
        expert_ids=expert_ids,
        similarity_matrix=similarity_matrix,
    )

    with ZipFile(args.output_zip, "w", ZIP_DEFLATED) as archive:
        archive.write(args.output_report, args.output_report.name)
        for path in args.output_dir.glob("*.csv"):
            archive.write(path, path.name)

    result = {
        "files_found": len(files),
        "unique_experts": len(expert_ids),
        "observations": len(long_rows),
        "missing_rating_cells": sum(row.get("missing_rating_cells", 0) for row in validations),
        "invalid_rating_cells": sum(row.get("invalid_rating_cells", 0) for row in validations),
        "duplicate_pairs": sum(1 for row in similarity if row["duplicate_flag"] == "YES"),
        "high_similarity_pairs": sum(1 for row in similarity if row["duplicate_flag"] == "HIGH_SIMILARITY"),
        "top_similarity_pairs": sorted(similarity, key=lambda row: row["both_usage_and_difficulty_same_percent"], reverse=True)[:5],
        "reliability": reliability,
        "report": str(args.output_report),
        "output_zip": str(args.output_zip),
    }
    print(json.dumps(result, indent=2))

    if args.extract_dir.exists():
        shutil.rmtree(args.extract_dir)

if __name__ == "__main__":
    main()
