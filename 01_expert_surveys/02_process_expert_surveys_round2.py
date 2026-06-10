#!/usr/bin/env python3
"""
Reproduce Delphi Round 2 expert survey analysis.

Inputs:
  - Round 1 completed expert survey ZIP (15 files, 209 skills each)
  - Round 2 completed expert survey ZIP (15 files, low-consensus skills only)

Outputs:
  - Excel analysis report
  - CSV outputs folder
  - ZIP archive of CSV outputs

The script parses linguistic dropdown labels such as "4 - Frequent / important"
by extracting the numeric prefix. It then combines Round 1 full data with Round 2
revisions for the targeted low-consensus skills and computes convergence and
reliability statistics.
"""

import argparse
import csv
import math
import os
import re
import shutil
import statistics
import tempfile
import zipfile
from collections import defaultdict, Counter
from itertools import combinations
from pathlib import Path

import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VALID_VALUES = {1, 2, 3, 4, 5}
NUMERIC_PREFIX_RE = re.compile(r"^\s*([1-5])")
EXPERT_RE = re.compile(r"(E\d{2}(?:_[A-Za-z0-9]+)?)")


def label_to_num(value):
    """Convert dropdown label or number to integer 1..5, else None."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if int(value) == value and int(value) in VALID_VALUES:
            return int(value)
        return None
    text = str(value).strip()
    m = NUMERIC_PREFIX_RE.match(text)
    if m:
        return int(m.group(1))
    return None


def safe_float(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def extract_zip(zip_path, out_dir):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as zf:
        names = [n for n in zf.namelist() if n.lower().endswith(".xlsx") and not os.path.basename(n).startswith("~$")]
        zf.extractall(out_dir)
    return [out_dir / n for n in names]


def infer_expert_id_from_filename(path):
    m = EXPERT_RE.search(Path(path).stem)
    return m.group(1) if m else Path(path).stem


def read_round1_file(path):
    wb = load_workbook(path, data_only=False)
    if "Survey" in wb.sheetnames:
        ws = wb["Survey"]
    else:
        ws = wb[wb.sheetnames[0]]

    # Expert metadata from the standard Round 1 template.
    expert_id = ws["A2"].value or infer_expert_id_from_filename(path)
    round_no = ws["B2"].value
    years_experience = ws["D2"].value
    lcdp_experience_count = ws["E2"].value
    missing_counter = ws["G2"].value
    completion_status = ws["H2"].value

    records = []
    invalid = []
    missing = []
    for row in range(5, ws.max_row + 1):
        skill_id = ws.cell(row, 3).value
        if skill_id is None:
            continue
        project_stage = ws.cell(row, 1).value
        project_task = ws.cell(row, 2).value
        skill_name = ws.cell(row, 4).value
        usage_label = ws.cell(row, 5).value
        difficulty_label = ws.cell(row, 6).value
        usage = label_to_num(usage_label)
        difficulty = label_to_num(difficulty_label)
        if usage is None:
            missing.append((skill_id, "Usage")) if usage_label in (None, "") else invalid.append((skill_id, "Usage", usage_label))
        if difficulty is None:
            missing.append((skill_id, "Difficulty")) if difficulty_label in (None, "") else invalid.append((skill_id, "Difficulty", difficulty_label))
        weight = usage * difficulty if usage is not None and difficulty is not None else None
        records.append({
            "expert_id": str(expert_id),
            "round": 1,
            "years_experience": years_experience,
            "lcdp_experience_count": lcdp_experience_count,
            "skill_id": int(skill_id),
            "project_stage": project_stage,
            "project_task": project_task,
            "skill_name": skill_name,
            "usage_label": usage_label,
            "usage": usage,
            "difficulty_label": difficulty_label,
            "difficulty": difficulty,
            "weight": weight,
            "source_file": Path(path).name,
        })
    validation = {
        "source_file": Path(path).name,
        "expert_id": str(expert_id),
        "round": round_no,
        "skill_rows": len(records),
        "missing_ratings": len(missing),
        "invalid_ratings": len(invalid),
        "completion_status_cell": completion_status,
        "years_experience": years_experience,
        "lcdp_experience_count": lcdp_experience_count,
    }
    return records, validation, missing, invalid


def read_round2_file(path):
    wb = load_workbook(path, data_only=False)
    if "Round 2 Survey" in wb.sheetnames:
        ws = wb["Round 2 Survey"]
    else:
        ws = wb[wb.sheetnames[0]]

    expert_id = ws["B2"].value or infer_expert_id_from_filename(path)
    round_no = ws["B3"].value
    years_experience = ws["B4"].value
    lcdp_experience_count = ws["B5"].value
    missing_counter = ws["B6"].value
    completion_status = ws["B7"].value

    records = []
    invalid = []
    missing = []
    for row in range(12, ws.max_row + 1):
        skill_id = ws.cell(row, 1).value
        if skill_id is None:
            continue
        project_stage = ws.cell(row, 2).value
        project_task = ws.cell(row, 3).value
        skill_name = ws.cell(row, 4).value
        r1_usage_label = ws.cell(row, 5).value
        panel_usage_mean = safe_float(ws.cell(row, 6).value)
        panel_usage_sd = safe_float(ws.cell(row, 7).value)
        r2_usage_label = ws.cell(row, 8).value
        r1_difficulty_label = ws.cell(row, 9).value
        panel_difficulty_mean = safe_float(ws.cell(row, 10).value)
        panel_difficulty_sd = safe_float(ws.cell(row, 11).value)
        r2_difficulty_label = ws.cell(row, 12).value
        r1_weight = safe_float(ws.cell(row, 13).value)
        panel_weight_mean = safe_float(ws.cell(row, 14).value)
        panel_weight_sd = safe_float(ws.cell(row, 15).value)
        comment = ws.cell(row, 16).value
        r2_usage = label_to_num(r2_usage_label)
        r2_difficulty = label_to_num(r2_difficulty_label)
        r1_usage = label_to_num(r1_usage_label)
        r1_difficulty = label_to_num(r1_difficulty_label)
        if r2_usage is None:
            missing.append((skill_id, "Round 2 Usage")) if r2_usage_label in (None, "") else invalid.append((skill_id, "Round 2 Usage", r2_usage_label))
        if r2_difficulty is None:
            missing.append((skill_id, "Round 2 Difficulty")) if r2_difficulty_label in (None, "") else invalid.append((skill_id, "Round 2 Difficulty", r2_difficulty_label))
        r2_weight = r2_usage * r2_difficulty if r2_usage is not None and r2_difficulty is not None else None
        records.append({
            "expert_id": str(expert_id),
            "round": 2,
            "years_experience": years_experience,
            "lcdp_experience_count": lcdp_experience_count,
            "skill_id": int(skill_id),
            "project_stage": project_stage,
            "project_task": project_task,
            "skill_name": skill_name,
            "round1_usage_label": r1_usage_label,
            "round1_usage": r1_usage,
            "panel_usage_mean_r1": panel_usage_mean,
            "panel_usage_sd_r1": panel_usage_sd,
            "round2_usage_label": r2_usage_label,
            "round2_usage": r2_usage,
            "round1_difficulty_label": r1_difficulty_label,
            "round1_difficulty": r1_difficulty,
            "panel_difficulty_mean_r1": panel_difficulty_mean,
            "panel_difficulty_sd_r1": panel_difficulty_sd,
            "round2_difficulty_label": r2_difficulty_label,
            "round2_difficulty": r2_difficulty,
            "round1_weight": r1_weight,
            "panel_weight_mean_r1": panel_weight_mean,
            "panel_weight_sd_r1": panel_weight_sd,
            "round2_weight": r2_weight,
            "comment": comment,
            "source_file": Path(path).name,
        })
    validation = {
        "source_file": Path(path).name,
        "expert_id": str(expert_id),
        "round": round_no,
        "skill_rows": len(records),
        "missing_ratings": len(missing),
        "invalid_ratings": len(invalid),
        "completion_status_cell": completion_status,
        "years_experience": years_experience,
        "lcdp_experience_count": lcdp_experience_count,
    }
    return records, validation, missing, invalid


def summarize_by_skill(records, usage_key="usage", difficulty_key="difficulty", weight_key="weight"):
    by_skill = defaultdict(list)
    for r in records:
        if r.get(usage_key) is None or r.get(difficulty_key) is None or r.get(weight_key) is None:
            continue
        by_skill[r["skill_id"]].append(r)

    rows = []
    for skill_id in sorted(by_skill):
        items = by_skill[skill_id]
        usage = [x[usage_key] for x in items]
        diff = [x[difficulty_key] for x in items]
        weight = [x[weight_key] for x in items]
        first = items[0]
        def sd(v): return statistics.stdev(v) if len(v) > 1 else 0.0
        row = {
            "skill_id": skill_id,
            "project_stage": first.get("project_stage"),
            "project_task": first.get("project_task"),
            "skill_name": first.get("skill_name"),
            "n_experts": len(items),
            "usage_mean": statistics.mean(usage),
            "usage_median": statistics.median(usage),
            "usage_sd": sd(usage),
            "usage_min": min(usage),
            "usage_max": max(usage),
            "difficulty_mean": statistics.mean(diff),
            "difficulty_median": statistics.median(diff),
            "difficulty_sd": sd(diff),
            "difficulty_min": min(diff),
            "difficulty_max": max(diff),
            "weight_mean": statistics.mean(weight),
            "weight_median": statistics.median(weight),
            "weight_sd": sd(weight),
            "weight_min": min(weight),
            "weight_max": max(weight),
            "weight_range": max(weight) - min(weight),
        }
        if row["weight_sd"] <= 3:
            row["consensus_level"] = "High"
        elif row["weight_sd"] <= 5:
            row["consensus_level"] = "Moderate"
        else:
            row["consensus_level"] = "Low"
        rows.append(row)
    return rows


def pivot_matrix(records, value_key):
    experts = sorted({r["expert_id"] for r in records})
    skills = sorted({r["skill_id"] for r in records})
    index = {(r["expert_id"], r["skill_id"]): r.get(value_key) for r in records}
    mat = np.array([[index.get((e, s), np.nan) for s in skills] for e in experts], dtype=float)
    return experts, skills, mat


def cronbach_alpha(mat):
    # rows=experts, cols=skills/items. Alpha across experts as raters over skill items.
    x = np.asarray(mat, dtype=float)
    x = x[~np.isnan(x).any(axis=1)]
    if x.shape[0] < 2 or x.shape[1] < 2:
        return float("nan")
    k = x.shape[0]
    item_vars = np.var(x, axis=1, ddof=1)
    total = np.sum(x, axis=0)
    total_var = np.var(total, ddof=1)
    if total_var == 0:
        return float("nan")
    return float((k / (k - 1)) * (1 - (np.sum(item_vars) / total_var)))


def icc_two_way_random_absolute(mat):
    # ICC(2,1) and ICC(2,k) using rows=targets/skills, cols=raters/experts.
    # Input mat rows experts x skills, transpose first.
    x = np.asarray(mat, dtype=float).T
    x = x[~np.isnan(x).any(axis=1)]
    n, k = x.shape  # targets, raters
    if n < 2 or k < 2:
        return float("nan"), float("nan")
    grand_mean = np.mean(x)
    mean_targets = np.mean(x, axis=1)
    mean_raters = np.mean(x, axis=0)
    ss_targets = k * np.sum((mean_targets - grand_mean) ** 2)
    ss_raters = n * np.sum((mean_raters - grand_mean) ** 2)
    ss_error = np.sum((x - mean_targets[:, None] - mean_raters[None, :] + grand_mean) ** 2)
    ms_targets = ss_targets / (n - 1)
    ms_raters = ss_raters / (k - 1)
    ms_error = ss_error / ((n - 1) * (k - 1))
    denom_single = ms_targets + (k - 1) * ms_error + (k * (ms_raters - ms_error) / n)
    icc_single = (ms_targets - ms_error) / denom_single if denom_single != 0 else float("nan")
    denom_avg = ms_targets + ((ms_raters - ms_error) / n)
    icc_avg = (ms_targets - ms_error) / denom_avg if denom_avg != 0 else float("nan")
    return float(icc_single), float(icc_avg)


def reliability_table(records, label):
    rows = []
    for value_key, measure in [("usage", "Usage Intensity"), ("difficulty", "Difficulty of Acquisition"), ("weight", "Synthetic Weight")]:
        experts, skills, mat = pivot_matrix(records, value_key)
        alpha = cronbach_alpha(mat)
        icc_single, icc_avg = icc_two_way_random_absolute(mat)
        rows.append({
            "dataset": label,
            "measure": measure,
            "n_experts": len(experts),
            "n_skills": len(skills),
            "cronbach_alpha": alpha,
            "icc_single_expert": icc_single,
            "icc_average_experts": icc_avg,
        })
    return rows


def pairwise_similarity(records, usage_key="usage", difficulty_key="difficulty"):
    by_exp = defaultdict(dict)
    for r in records:
        by_exp[r["expert_id"]][r["skill_id"]] = (r.get(usage_key), r.get(difficulty_key))
    rows = []
    experts = sorted(by_exp)
    for a, b in combinations(experts, 2):
        common = sorted(set(by_exp[a]) & set(by_exp[b]))
        full_matches = sum(1 for s in common if by_exp[a][s] == by_exp[b][s])
        usage_matches = sum(1 for s in common if by_exp[a][s][0] == by_exp[b][s][0])
        diff_matches = sum(1 for s in common if by_exp[a][s][1] == by_exp[b][s][1])
        rows.append({
            "expert_a": a,
            "expert_b": b,
            "n_common_skills": len(common),
            "full_answer_matches": full_matches,
            "full_similarity_pct": full_matches / len(common) if common else None,
            "usage_matches": usage_matches,
            "usage_similarity_pct": usage_matches / len(common) if common else None,
            "difficulty_matches": diff_matches,
            "difficulty_similarity_pct": diff_matches / len(common) if common else None,
        })
    rows.sort(key=lambda x: x["full_similarity_pct"] if x["full_similarity_pct"] is not None else -1, reverse=True)
    return rows


def distributions(records, usage_key="usage", difficulty_key="difficulty", weight_key="weight"):
    rows = []
    for name, key in [("Usage Intensity", usage_key), ("Difficulty of Acquisition", difficulty_key), ("Synthetic Weight", weight_key)]:
        cnt = Counter(r.get(key) for r in records if r.get(key) is not None)
        total = sum(cnt.values())
        for value in sorted(cnt):
            rows.append({"measure": name, "value": value, "count": cnt[value], "proportion": cnt[value] / total if total else None})
    return rows


def write_csv(path, rows, fieldnames=None):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col[:1000])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)


def add_sheet_from_rows(wb, name, rows, fieldnames=None):
    ws = wb.create_sheet(name[:31])
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(f) for f in fieldnames])
    style_sheet(ws)
    return ws


def make_summary(round1_records, round2_records, final_records, val_r1, val_r2, convergence_rows, rel_rows, pair_rows):
    r2_skill_ids = {r["skill_id"] for r in round2_records}
    improved = sum(1 for r in convergence_rows if r.get("weight_sd_change") is not None and r["weight_sd_change"] < 0)
    unchanged_or_worse = len(convergence_rows) - improved
    avg_change = statistics.mean([r["weight_sd_change"] for r in convergence_rows if r.get("weight_sd_change") is not None]) if convergence_rows else None
    highest_pair = pair_rows[0] if pair_rows else {}
    return [
        {"item": "Round 1 experts", "value": len({r["expert_id"] for r in round1_records})},
        {"item": "Round 1 skills", "value": len({r["skill_id"] for r in round1_records})},
        {"item": "Round 1 observations", "value": len(round1_records)},
        {"item": "Round 2 experts", "value": len({r["expert_id"] for r in round2_records})},
        {"item": "Round 2 targeted skills", "value": len(r2_skill_ids)},
        {"item": "Round 2 observations", "value": len(round2_records)},
        {"item": "Round 2 files with missing ratings", "value": sum(1 for v in val_r2 if v["missing_ratings"] > 0)},
        {"item": "Round 2 files with invalid ratings", "value": sum(1 for v in val_r2 if v["invalid_ratings"] > 0)},
        {"item": "Final observations after Round 2 substitution", "value": len(final_records)},
        {"item": "Low-consensus skills with reduced synthetic-weight SD", "value": improved},
        {"item": "Low-consensus skills without reduced synthetic-weight SD", "value": unchanged_or_worse},
        {"item": "Mean change in synthetic-weight SD (Round 2 - Round 1)", "value": avg_change},
        {"item": "Highest final pairwise similarity pair", "value": f"{highest_pair.get('expert_a','')} vs {highest_pair.get('expert_b','')}"},
        {"item": "Highest final full-answer similarity", "value": highest_pair.get("full_similarity_pct")},
    ]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--round1-zip", default=str(Path(__file__).parent / "input" / "completed_expert_surveys_round1_FINAL.zip"), help="ZIP with 15 completed Round 1 Excel files")
    parser.add_argument("--round2-zip", default=str(Path(__file__).parent / "input" / "completed_expert_surveys_round2_FINAL.zip"), help="ZIP with 15 completed Round 2 Excel files")
    parser.add_argument("--output-report", default=str(Path(__file__).parent / "output" / "expert_survey_round2_analysis_report.xlsx"), help="Output Excel analysis report path")
    parser.add_argument("--output-dir", default=str(Path(__file__).parent / "output"), help="Output directory for CSV files")
    parser.add_argument("--output-zip", default=str(Path(__file__).parent / "output" / "expert_survey_round2_outputs.zip"), help="Output ZIP archive for CSV files")
    args = parser.parse_args()

    out_dir = Path(args.output_dir)
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmp:
        r1_dir = Path(tmp) / "round1"
        r2_dir = Path(tmp) / "round2"
        r1_files = extract_zip(args.round1_zip, r1_dir)
        r2_files = extract_zip(args.round2_zip, r2_dir)

        round1_records, val_r1, missing_r1, invalid_r1 = [], [], [], []
        for f in sorted(r1_files):
            recs, val, miss, inv = read_round1_file(f)
            round1_records.extend(recs)
            val_r1.append(val)
            missing_r1.extend([{"source_file": Path(f).name, "skill_id": m[0], "field": m[1]} for m in miss])
            invalid_r1.extend([{"source_file": Path(f).name, "skill_id": i[0], "field": i[1], "value": i[2]} for i in inv])

        round2_records, val_r2, missing_r2, invalid_r2 = [], [], [], []
        for f in sorted(r2_files):
            recs, val, miss, inv = read_round2_file(f)
            round2_records.extend(recs)
            val_r2.append(val)
            missing_r2.extend([{"source_file": Path(f).name, "skill_id": m[0], "field": m[1]} for m in miss])
            invalid_r2.extend([{"source_file": Path(f).name, "skill_id": i[0], "field": i[1], "value": i[2]} for i in inv])

    # Round 1 targeted subset and Round 2 records by expert/skill.
    r2_skill_ids = sorted({r["skill_id"] for r in round2_records})
    r2_map = {(r["expert_id"], r["skill_id"]): r for r in round2_records}

    # Build final records: Round 1 all skills, with Round 2 values replacing targeted low-consensus skills.
    final_records = []
    round1_targeted_records = []
    for r in round1_records:
        f = dict(r)
        f["final_round_source"] = "Round 1"
        if r["skill_id"] in r2_skill_ids:
            round1_targeted_records.append(r)
            rr = r2_map.get((r["expert_id"], r["skill_id"]))
            if rr:
                f["usage_label"] = rr["round2_usage_label"]
                f["usage"] = rr["round2_usage"]
                f["difficulty_label"] = rr["round2_difficulty_label"]
                f["difficulty"] = rr["round2_difficulty"]
                f["weight"] = rr["round2_weight"]
                f["final_round_source"] = "Round 2"
        final_records.append(f)

    # Summaries.
    r1_target_consensus = summarize_by_skill(round1_targeted_records)
    r2_consensus = summarize_by_skill(
        [{
            "expert_id": r["expert_id"], "skill_id": r["skill_id"], "project_stage": r["project_stage"],
            "project_task": r["project_task"], "skill_name": r["skill_name"],
            "usage": r["round2_usage"], "difficulty": r["round2_difficulty"], "weight": r["round2_weight"],
        } for r in round2_records]
    )
    final_consensus = summarize_by_skill(final_records)

    r1_target_by_skill = {r["skill_id"]: r for r in r1_target_consensus}
    r2_by_skill = {r["skill_id"]: r for r in r2_consensus}
    convergence_rows = []
    for sid in r2_skill_ids:
        a = r1_target_by_skill.get(sid, {})
        b = r2_by_skill.get(sid, {})
        row = {
            "skill_id": sid,
            "project_stage": b.get("project_stage") or a.get("project_stage"),
            "project_task": b.get("project_task") or a.get("project_task"),
            "skill_name": b.get("skill_name") or a.get("skill_name"),
            "round1_usage_mean": a.get("usage_mean"),
            "round2_usage_mean": b.get("usage_mean"),
            "usage_mean_change": (b.get("usage_mean") - a.get("usage_mean")) if a.get("usage_mean") is not None and b.get("usage_mean") is not None else None,
            "round1_usage_sd": a.get("usage_sd"),
            "round2_usage_sd": b.get("usage_sd"),
            "usage_sd_change": (b.get("usage_sd") - a.get("usage_sd")) if a.get("usage_sd") is not None and b.get("usage_sd") is not None else None,
            "round1_difficulty_mean": a.get("difficulty_mean"),
            "round2_difficulty_mean": b.get("difficulty_mean"),
            "difficulty_mean_change": (b.get("difficulty_mean") - a.get("difficulty_mean")) if a.get("difficulty_mean") is not None and b.get("difficulty_mean") is not None else None,
            "round1_difficulty_sd": a.get("difficulty_sd"),
            "round2_difficulty_sd": b.get("difficulty_sd"),
            "difficulty_sd_change": (b.get("difficulty_sd") - a.get("difficulty_sd")) if a.get("difficulty_sd") is not None and b.get("difficulty_sd") is not None else None,
            "round1_weight_mean": a.get("weight_mean"),
            "round2_weight_mean": b.get("weight_mean"),
            "weight_mean_change": (b.get("weight_mean") - a.get("weight_mean")) if a.get("weight_mean") is not None and b.get("weight_mean") is not None else None,
            "round1_weight_sd": a.get("weight_sd"),
            "round2_weight_sd": b.get("weight_sd"),
            "weight_sd_change": (b.get("weight_sd") - a.get("weight_sd")) if a.get("weight_sd") is not None and b.get("weight_sd") is not None else None,
        }
        row["weight_sd_reduced"] = row["weight_sd_change"] is not None and row["weight_sd_change"] < 0
        convergence_rows.append(row)

    reliability_rows = []
    reliability_rows.extend(reliability_table(round1_records, "Round 1 all 209 skills"))
    reliability_rows.extend(reliability_table(round1_targeted_records, "Round 1 targeted 46 skills"))
    reliability_rows.extend(reliability_table([{
        "expert_id": r["expert_id"], "skill_id": r["skill_id"], "usage": r["round2_usage"],
        "difficulty": r["round2_difficulty"], "weight": r["round2_weight"],
    } for r in round2_records], "Round 2 targeted 46 skills"))
    reliability_rows.extend(reliability_table(final_records, "Final after Round 2 substitution"))

    pair_final = pairwise_similarity(final_records)
    pair_round2 = pairwise_similarity([{
        "expert_id": r["expert_id"], "skill_id": r["skill_id"], "usage": r["round2_usage"],
        "difficulty": r["round2_difficulty"],
    } for r in round2_records])
    dist_final = distributions(final_records)
    dist_round2 = distributions([{
        "usage": r["round2_usage"], "difficulty": r["round2_difficulty"], "weight": r["round2_weight"]
    } for r in round2_records])

    summary_rows = make_summary(round1_records, round2_records, final_records, val_r1, val_r2, convergence_rows, reliability_rows, pair_final)

    # Write CSV files.
    write_csv(out_dir / "round2_validation_summary.csv", val_r2)
    write_csv(out_dir / "round2_long_format.csv", round2_records)
    write_csv(out_dir / "round2_convergence_by_skill.csv", convergence_rows)
    write_csv(out_dir / "final_expert_survey_long_format_after_round2.csv", final_records)
    write_csv(out_dir / "final_consensus_weights_after_round2.csv", final_consensus)
    write_csv(out_dir / "expert_reliability_round2_and_final.csv", reliability_rows)
    write_csv(out_dir / "expert_pairwise_similarity_final_after_round2.csv", pair_final)
    write_csv(out_dir / "expert_pairwise_similarity_round2_targeted.csv", pair_round2)
    write_csv(out_dir / "rating_distributions_final_after_round2.csv", dist_final)
    write_csv(out_dir / "rating_distributions_round2_targeted.csv", dist_round2)
    write_csv(out_dir / "round2_missing_ratings.csv", missing_r2, ["source_file", "skill_id", "field"])
    write_csv(out_dir / "round2_invalid_ratings.csv", invalid_r2, ["source_file", "skill_id", "field", "value"])

    # Excel report.
    wb = Workbook()
    wb.remove(wb.active)
    add_sheet_from_rows(wb, "Summary", summary_rows)
    add_sheet_from_rows(wb, "Round2Validation", val_r2)
    add_sheet_from_rows(wb, "Reliability", reliability_rows)
    add_sheet_from_rows(wb, "ConvergenceBySkill", convergence_rows)
    add_sheet_from_rows(wb, "FinalConsensusWeights", final_consensus)
    add_sheet_from_rows(wb, "Round2LongData", round2_records)
    add_sheet_from_rows(wb, "FinalLongData", final_records)
    add_sheet_from_rows(wb, "PairwiseSimilarityFinal", pair_final)
    add_sheet_from_rows(wb, "PairwiseSimilarityR2", pair_round2)
    add_sheet_from_rows(wb, "DistributionsFinal", dist_final)
    add_sheet_from_rows(wb, "DistributionsR2", dist_round2)
    if missing_r2:
        add_sheet_from_rows(wb, "MissingRatings", missing_r2)
    if invalid_r2:
        add_sheet_from_rows(wb, "InvalidRatings", invalid_r2)
    Path(args.output_report).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output_report)

    # Zip CSV outputs.
    if os.path.exists(args.output_zip):
        os.remove(args.output_zip)
    with zipfile.ZipFile(args.output_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for f in sorted(out_dir.glob("*.csv")):
            zf.write(f, arcname=f.name)

    # Brief console summary.
    print("Round 2 analysis complete")
    for row in summary_rows:
        print(f"{row['item']}: {row['value']}")

    if args.extract_dir.exists():
        import shutil
        shutil.rmtree(args.extract_dir)

if __name__ == "__main__":
    main()
